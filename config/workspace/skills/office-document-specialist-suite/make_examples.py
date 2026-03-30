#!/usr/bin/env python3
"""
예시 파일 생성 스크립트
setup-agent.sh 실행 시 호출되어 예시 파일을 생성한다
"""
# ref: https://python-pptx.readthedocs.io / https://openpyxl.readthedocs.io
import os
import json
import sys

def make_example_ppt(skills_dir, output_dir):
    sys.path.insert(0, skills_dir)
    from make_ppt import make_ppt

    example = {
        "title": "AI 트렌드 2025",
        "subtitle": "Clari AI 예시 발표자료",
        "slides": [
            {"title": "LLM 발전", "content": "• 대규모 언어 모델 급속 발전\n• GPT-4, Claude, Gemini 경쟁 심화\n• 오픈소스 모델 부상 (Llama, Qwen 등)"},
            {"title": "AI 에이전트", "content": "• 자율적 작업 수행 에이전트 확산\n• 멀티에이전트 협업 시스템\n• 실제 업무 자동화 본격화"},
            {"title": "멀티모달 AI", "content": "• 텍스트, 이미지, 음성 통합 처리\n• 영상 이해 및 생성 능력 향상\n• 실시간 처리 성능 개선"},
            {"title": "결론", "content": "• AI 기술의 실용화 가속\n• 기업 도입률 급증\n• 윤리적 AI 규제 논의 본격화"},
        ]
    }

    for theme in ["dark", "light"]:
        output = os.path.join(output_dir, f"example_{theme}.pptx")
        make_ppt(json.dumps(example), output, theme)
        print(f"예시 PPT 생성 완료: {output}")

def make_example_excel(output_dir):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "월별 매출"

    headers = ["월", "매출(만원)", "비용(만원)", "수익(만원)", "수익률(%)"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="7C3AED")
        cell.alignment = Alignment(horizontal="center")

    months = ["1월","2월","3월","4월","5월","6월","7월","8월","9월","10월","11월","12월"]
    data = [
        (1000, 600), (1200, 700), (1500, 850), (1300, 750),
        (1800, 1000), (2000, 1100), (1900, 1050), (2200, 1200),
        (2500, 1350), (2300, 1250), (2800, 1500), (3000, 1600)
    ]
    for i, (month, (revenue, cost)) in enumerate(zip(months, data), 2):
        profit = revenue - cost
        rate = round(profit / revenue * 100, 1)
        ws.cell(row=i, column=1, value=month)
        ws.cell(row=i, column=2, value=revenue)
        ws.cell(row=i, column=3, value=cost)
        ws.cell(row=i, column=4, value=profit)
        ws.cell(row=i, column=5, value=rate)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15

    output = os.path.join(output_dir, "example_excel.xlsx")
    wb.save(output)
    print(f"예시 Excel 생성 완료: {output}")

if __name__ == "__main__":
    skills_dir = os.path.join(
        os.environ.get("OPENCLAW_WORKSPACE", os.path.expanduser("~/.openclaw/workspace")),
        "skills", "office-document-specialist-suite"
    )
    output_dir = os.environ.get("OUTPUT_DIR", "/workspace/work")
    os.makedirs(output_dir, exist_ok=True)
    make_example_ppt(skills_dir, output_dir)
    make_example_excel(output_dir)
    print("모든 예시 파일 생성 완료")
